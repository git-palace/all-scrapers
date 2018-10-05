# from __future__ import unicode_literals
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from lxml import etree

from lxml import html

import pdb

import csv

import re

from openpyxl import load_workbook



class business_email_list_xlsx(scrapy.Spider):

	name = 'business-email-list-xlsx'

	domain = 'yelp.com'

	history = []

	business_list = []

	count = 0


	def __init__(self):

		# self.wb = load_workbook('1.xlsx')
		self.wb = load_workbook('00_business-list_20180928.xlsx')

		self.ws = self.wb.active


	
	def start_requests(self):

		yield scrapy.Request(url='https://yelp.com', callback=self.start_business_emails)



	def start_business_emails(self, response):

		r_idx = 1

		for row in self.ws.iter_rows():

			if r_idx != 1:

				business_name = self.ws.cell(r_idx, 1).value

				url = self.ws.cell(r_idx, 2).value

				if 'http' in url:

					url = url

				else:

					url = 'http://' + url

				yield scrapy.Request(url=url, callback=self.parse, meta={'business_name': business_name}, dont_filter=True)

			r_idx += 1



	def parse(self, response):

		website = response.url

		if 'website' in response.meta:

			website = response.meta['website']

		item = ChainItem()

		item['business_name'] = response.meta['business_name']

		item['website'] = website

		
		text_arr = response.xpath('//a[starts-with(@href, "mail")]/@href').extract()

		if not len(text_arr):

			text_arr = response.xpath('//body//text()').extract()

		email = ''

		for text in text_arr:

			email = self.check_email(text)

			if email:

				break

		if email:

			self.count += 1

			print '======================================================='
			print '======================================================='
			print '======================== %s ==========================' % (self.count)
			print '=========== %s, %s ==========' % (website, email)
			print '======================================================='
			print '======================================================='

			item['email'] = email

			yield item

		else:

			contact_link = response.xpath('//a[contains(text(), "contact") or contains(text(), "Contact") or contains(text(), "CONTACT") or contains(@href, "contact") or contains(@href, "Contact") or contains(@href, "CONTACT")]/@href').extract_first()

			if contact_link:

				url = '%s%s' % (website+'/' if 'http' not in contact_link else '', contact_link)

				url = url.replace('//'+contact_link, '/'+contact_link)

				if len(url.split('//')) > 2:

					url = '/'.join(['//'.join(url.split('//')[:2])]+url.split('//')[2:])

				if url not in self.history:

					self.history.append(url)

					yield scrapy.Request(url=url, callback=self.parse, meta={'business_name': response.meta['business_name'], 'website': website})

				else:

					yield item

			else:

				yield item



	def check_email(self, text):

		text = text.replace('mailto:', '')

		for text in self.validate(text).split():

			if '@' in text:

				match = re.search(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)", text)

				return match.group(0) if match else False

		return False



	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').strip()

		except:

			pass


	def eliminate_space(self, items):

		tmp = []

		for item in items:

			if self.validate(item) != '':

				tmp.append(self.validate(item))

		return tmp