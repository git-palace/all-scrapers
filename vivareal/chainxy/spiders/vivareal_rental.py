# from __future__ import unicode_literals
# -*- coding: utf-8 -*-
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from lxml import etree

from lxml import html

import pdb

from openpyxl import Workbook 

import random




class vivareal_rental(scrapy.Spider):

	name = 'vivareal_rental'

	domain = ''

	history = []

	workbook_list = []

	location_list = []

	businessType = 'RENTAL'

	proxy_list = []

	sheet_headers = [
		'Title',
		'Business Type',
		'Address',
		'Area',
		'Bedrooms',
		'Bathrooms',
		'Parkings',
		'Price',
		'Condo Price',
		'Code',
		'Property Type',
		'Phones',
		'Description',
		'Amenities',
		'Images',
		'IPTU Price',
		'Value with Condominum'
	]

	file_idx = 0

	total_scraped_count = 0


	def __init__(self):

		with open('require/neighborhoods.json', 'r') as f:

			self.location_list = json.load(f)

		with open('require/proxy_ips_%s.json' % (self.businessType), 'r') as f:

			self.proxy_list = json.load(f)

		self.create_file()

		pass

	
	def start_requests(self):

		for location in self.location_list:

			for neighborhood in location['neighborhoods']:

				addressNeighborhood = neighborhood

				org_url = 'https://glue-api.vivareal.com/v1/listings?addressCity=%s&addressCountry=BR&addressState=%sfilterPricingInfoBusinessType=%s&facets=amenities&filter=(address.neighborhood:"%s") AND pricingInfos.businessType:"%s" AND propertyType:"UNIT" AND listingType:"USED"&filterListingType=USED&includeFields=addresses,listingsLocation,seo,search,url,expansion,nearby,developments&filterPropertyType=UNIT&developmentsSize=5' % (location['city'], location['state'], self.businessType, neighborhood, self.businessType)

				url = '%s&size=100&from=0' % (org_url)

				# url  = 'https://glue-api.vivareal.com/v1/listings?size=100&from=0'

				proxy = 'http://%s' % ( random.choice(self.proxy_list))

				yield scrapy.Request(url=url, callback=self.parse, meta={'org_url': org_url, 'paged': 0, 'proxy': proxy})




	def parse(self, response):

		paged = response.meta['paged'] + 1

		json_data = json.loads(response.body)

		if len(json_data['search']['result']['listings']):

			for data in json_data['search']['result']['listings']:

				try:

					listing = data['listing']

					if listing['id'] not in self.history:

						self.history.append(listing['id'])

						pricingInfos = data['listing']['pricingInfos'][0]

						listing_address = ''

						if 'address' in listing:

							listing_address = '%s, %s%s%s, %s - %s' % (
								listing['address']['street'] if 'street' in listing['address'] else '',
								listing['address']['streetNumber'] if 'streetNumber' in listing['address'] else '',
								' - ' if 'streetNumber' in listing['address'] else '',
								listing['address']['neighborhood'] if 'neighborhood' in listing['address'] else '',
								listing['address']['city'] if 'city' in listing['address'] else '',
								listing['address']['state'] if 'state' in listing['address'] else ''
							)

						item = {
							'Title': listing['title'] if 'title' in listing else None,
							'Business Type': pricingInfos['businessType'],
							'Address': listing_address,
							'Area': listing['usableAreas'][0] if len(listing['usableAreas']) else None,
							'Bedrooms': listing['bedrooms'][0] if len(listing['bedrooms']) else None,
							'Bathrooms': listing['bathrooms'][0] if len(listing['bathrooms']) else None,
							'Parkings': listing['parkingSpaces'][0] if len(listing['parkingSpaces']) else None,
							'Price': pricingInfos['price'] if 'price' in pricingInfos else None,
							'Condo Price': pricingInfos['monthlyCondoFee'] if 'monthlyCondoFee' in pricingInfos else '',
							'Code': listing['externalId'] if 'externalId' in listing else None,
							'Property Type': '\n'.join(listing['unitTypes'] if 'unitTypes' in listing else []),
							'Phones': '\n'.join(listing['contact']['phones'] if 'phones' in listing['contact'] else []),
							'Description': listing['description'].replace('<br>', '\n') if 'description' in listing else None,
							'Amenities': '\n'.join(listing['amenities'] if 'amenities' in listing else []).replace('_', ' '),
							'Images': '\n'.join(self.validate_images(listing['images'])) if 'images' in listing else None,
						}

						if self.businessType == 'RENTAL':

							item['IPTU Price'] = pricingInfos['yearlyIptu'] if 'yearlyIptu' in pricingInfos else ''

							if 'rentalTotalPrice' in pricingInfos and 'period' in pricingInfos:
									
								item['Value with Condominum'] = ('%s / %s') % (pricingInfos['rentalTotalPrice'], pricingInfos['period'])

						main_table_sheet = self.workbook_list[self.file_idx]['workbook'].active

						try:

							write_flag = True

							for key, value in item.items():

								try:

									main_table_sheet.cell( column = self.sheet_headers.index(key) + 1, row = self.workbook_list[self.file_idx]['cur_rowIndex'], value = value)

								except:

									write_flag = False

									pass

							if write_flag:

								for img_url in self.validate_images(listing['images']):

									yield scrapy.Request(url=img_url, callback=self.download_image, meta={'file_idx': self.file_idx})

								self.workbook_list[self.file_idx]['cur_rowIndex'] += 1

								self.workbook_list[self.file_idx]['workbook'].save( filename = self.workbook_list[self.file_idx]['filename'] )

								self.total_scraped_count += 1

								if self.total_scraped_count >= 3000 * (self.file_idx + 1):

									self.file_idx += 1

									self.create_file()
							
								print '========================================================='
								print '======================= %s =====================' % (self.total_scraped_count)
								print '========================================================='

						except:

							pass

				except:

					pass



			proxy = 'http://%s' % ( random.choice(self.proxy_list))

			yield scrapy.Request(url='%s&size=100&from=%s' % (response.meta['org_url'], paged*100), callback=self.parse, meta={'org_url': response.meta['org_url'], 'paged': paged, 'proxy': proxy})



	def validate_images(self, img_url_list):

		r_img_url_list = []

		for image_url in img_url_list:

			url = image_url.replace('{action}/{width}x{height}', 'fit-in/870x653')

			r_img_url_list.append(url)

		return r_img_url_list


	def download_image(self, response):

		dir_path = 'scraped-results/images/Scraping Vivareal %s' % (response.meta['file_idx']+1)

		if not os.path.exists(dir_path):
			
			os.makedirs(dir_path)
		
		name = dir_path + '/' + response.url.split('/')[-1]

		with open(name, 'wb') as f:

			f.write(response.body)


	def create_file(self):

		dir_path = 'scraped-results/sheets/Scraping Vivareal %s' % (self.file_idx+1)

		if not os.path.exists(dir_path):

			os.makedirs(dir_path)

		workbook = Workbook()

		main_table_sheet = workbook.active

		main_table_sheet.title = 'Scraped Results'

		for index, header in enumerate(self.sheet_headers):

			main_table_sheet.cell( column = index+1, row = 1, value = header )

		output_file_name = os.path.join(dir_path, 'scraped_data.xlsx')

		os.remove(output_file_name) if os.path.exists(output_file_name) else None

		self.workbook_list.append({
			'filename': output_file_name,
			'workbook': workbook,
			'cur_rowIndex': 2
		})


	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').strip()

		except:

			pass


	def eliminate_space(self, items):

	    tmp = []

	    for item in items:

	        if self.validate(item) != '':

	            tmp.append(self.validate(item))

	    return tmp