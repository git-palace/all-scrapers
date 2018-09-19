# from __future__ import unicode_literals
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from lxml import etree

from lxml import html

import pdb

from openpyxl import Workbook 




class vivareal(scrapy.Spider):

	name = 'vivareal'

	domain = ''

	history = []

	workbook_list = []

	sheet_headers = [
		'Title',
		'Code',
		'Business Type',
		'Price',
		'Condo Price',
		'IPTU Price',
		'Rental Total Price',
		'Rental Period',
		'Property Type',
		'Area',
		'Bedrooms',
		'Bathrooms',
		'Parkings',
		'Phones',
		'Description',
		'Amenities',
		'Images',
		'Publisher Name',
		'Publisher Street',
		'Publisher StreetNumber',
		'Publisher City',
		'Publisher State',
		'Publisher ZipCode',
		'Publisher District',
		'Publisher UnitNumber',
		'Publisher Neighborhood',
		'Publisher Zone'
	]

	file_idx = 0

	total_scraped_count = 0


	def __init__(self):

		self.create_file()

		pass

	
	def start_requests(self):

		url  = 'https://glue-api.vivareal.com/v1/listings?size=100&from=0'

		yield scrapy.Request(url=url, callback=self.parse, meta={'paged': 0})



	def parse(self, response):

		paged = response.meta['paged'] + 1

		json_data = json.loads(response.body)

		if len(json_data['search']['result']['listings']):

			for data in json_data['search']['result']['listings']:

				try:

					listing = data['listing']
					pricingInfos = data['listing']['pricingInfos'][0]

					item = {
						'Title': listing['title'],
						'Code': listing['externalId'],
						'Business Type': pricingInfos['businessType'],
						'Price': pricingInfos['price'],
						'Condo Price': pricingInfos['monthlyCondoFee'] if 'monthlyCondoFee' in pricingInfos else '',
						'IPTU Price': pricingInfos['yearlyIptu'] if 'yearlyIptu' in pricingInfos else '',
						'Rental Total Price': pricingInfos['rentalTotalPrice'] if 'rentalTotalPrice' in pricingInfos else '',
						'Rental Period': pricingInfos['period'] if 'period' in pricingInfos else '',
						'Property Type': '\n'.join(listing['unitTypes']),
						'Area': listing['totalAreas'][0],
						'Bedrooms': listing['bedrooms'][0],
						'Bathrooms': listing['bathrooms'][0],
						'Parkings': listing['parkingSpaces'][0],
						'Phones': '\n'.join(listing['contact']['phones']),
						'Description': listing['description'].replace('<br>', '\n'),
						'Amenities': '\n'.join(listing['amenities']).replace('_', ' '),
						'Images': '\n'.join(self.validate_images(listing['images'])),
						'Publisher Name': data['publisher']['name']
					}

					if 'address' in data['publisher']:
						pub_address = data['publisher']['address']

						item['Publisher Street'] = pub_address['street']
						item['Publisher StreetNumber'] = pub_address['streetNumber']
						item['Publisher City'] = pub_address['city']
						item['Publisher State'] = pub_address['state']
						item['Publisher ZipCode'] = pub_address['zipCode']
						item['Publisher District'] = pub_address['district']
						item['Publisher UnitNumber'] = pub_address['unitNumber']
						item['Publisher Neighborhood'] = pub_address['neighborhood']
						item['Publisher Zone'] = pub_address['zone']

					main_table_sheet = self.workbook_list[self.file_idx]['workbook'].active

					try:

						write_flag = True

						for key, value in item.items():

							try:

								main_table_sheet.cell( column = self.sheet_headers.index(key) + 1, row = self.workbook_list[self.file_idx]['cur_rowIndex'], value = value)

							except:

								write_flag = False

								pass

						if write_flag:

							# for img_url in self.validate_images(listing['images']):

								# yield scrapy.Request(url=img_url, callback=self.download_image, meta={'file_idx': self.file_idx})

							self.workbook_list[self.file_idx]['cur_rowIndex'] += 1

							self.workbook_list[self.file_idx]['workbook'].save( filename = self.workbook_list[self.file_idx]['filename'] )

							self.total_scraped_count += 1

							if self.total_scraped_count >= 5000 * (self.file_idx + 1):

								self.file_idx += 1

								self.create_file()
						
							print '========================================================='
							print '======================= %s =====================' % (self.total_scraped_count)
							print '========================================================='

					except:

						pass

				except:

					pass


			yield scrapy.Request(url='https://glue-api.vivareal.com/v1/listings?size=100&from=%s' % (paged*100), callback=self.parse, meta={'paged': paged})


	def validate_images(self, img_url_list):

		r_img_url_list = []

		for image_url in img_url_list:

			url = image_url.replace('{action}/{width}x{height}', 'fit-in/870x653')

			r_img_url_list.append(url)

		return r_img_url_list


	def download_image(self, response):

		dir_path = 'scraped-results/images/%s' % (response.meta['file_idx'])

		if not os.path.exists(dir_path):
			
			os.makedirs(dir_path)
		
		name = dir_path + '/' + response.url.split('/')[-1]

		with open(name, 'wb') as f:

			f.write(response.body)


	def create_file(self):

		dir_path = 'scraped-results/sheets/%s' % (self.file_idx)

		if not os.path.exists(dir_path):

			os.makedirs(dir_path)

		workbook = Workbook()

		main_table_sheet = workbook.active

		main_table_sheet.title = 'Scraped Results'

		for index, header in enumerate(self.sheet_headers):

			main_table_sheet.cell( column = index+1, row = 1, value = header )

		output_file_name = os.path.join(dir_path, 'scraped_data.xlsx')

		os.remove(output_file_name) if os.path.exists(output_file_name) else None

		self.workbook_list.append({
			'filename': output_file_name,
			'workbook': workbook,
			'cur_rowIndex': 2
		})


	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').strip()

		except:

			pass


	def eliminate_space(self, items):

	    tmp = []

	    for item in items:

	        if self.validate(item) != '':

	            tmp.append(self.validate(item))

	    return tmp