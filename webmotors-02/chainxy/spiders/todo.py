# -*- coding: utf-8 -*-

import scrapy
import json
import os
from scrapy.spiders import Spider
from scrapy.http import FormRequest
from scrapy.http import Request
from chainxy.items import ChainItem
from lxml import etree
from lxml import html
import pdb
from openpyxl import Workbook 


class todo(scrapy.Spider):
	s_count = 0;
	name = 'webmotors'
	domain = 'www.webmotors.com.br'
	history = []
	headers = {

		"accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
		"accept-encoding": "gzip, deflate, br",
		"pragma": "no-cache",
		"upgrade-insecure-requests": "1",
		"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36"
	}

	workbook_list = {}
	main_table_headers = [
		'make_model',
		'version_of_year',
		'financing',
		'city',
		'a_type',
		'price',
		'store_name',
		'phone_numbers',
		'year',
		'mileage',
		'exchange',
		'fuel',
		'color',
		'bodywork',
		'end_plate',
		'items',
		'seller_notes',
		'images'
	]

	def __init__(self):
		pass


	# validate
	def validate(self, item):
		try:
			
			return item.replace('\n', '').replace('\t', '').replace('\r', '').strip().encode('utf-8')

		except:

			pass


	# start scrapy
	def start_requests(self):

		brand_list = {
			'carro': '{"Principal":[{"N":"CHEVROLET","C":50804},{"N":"CITROËN","C":11604},{"N":"FIAT","C":49217},{"N":"FORD","C":34977},{"N":"HONDA","C":19421},{"N":"HYUNDAI","C":22760},{"N":"MITSUBISHI","C":10466},{"N":"NISSAN","C":8802},{"N":"PEUGEOT","C":10472},{"N":"RENAULT","C":25109},{"N":"TOYOTA","C":20028},{"N":"VOLKSWAGEN","C":51967},{"N":"VOLVO","C":2317}],"Common":[{"N":"AGRALE","C":1},{"N":"ALFA ROMEO","C":79},{"N":"AMAZONAS","C":0},{"N":"AMERICAR","C":3},{"N":"ASIA","C":3},{"N":"ASTON MARTIN","C":25},{"N":"AUDI","C":6262},{"N":"BEACH","C":1},{"N":"BENTLEY","C":15},{"N":"BIANCO","C":1},{"N":"BMW","C":8366},{"N":"BRM","C":28},{"N":"BUGRE","C":4},{"N":"BUGWAY","C":12},{"N":"BUICK","C":3},{"N":"CADILLAC","C":68},{"N":"CBP","C":1},{"N":"CBT","C":5},{"N":"CHAMONIX","C":6},{"N":"CHANA","C":11},{"N":"CHERY","C":664},{"N":"CHEVROLET","C":50804},{"N":"CHRYSLER","C":517},{"N":"CITROËN","C":11604},{"N":"DACON","C":3},{"N":"DAEWOO","C":3},{"N":"DAIHATSU","C":8},{"N":"DATSUN","C":1},{"N":"DE SOTO","C":1},{"N":"DKW-VEMAG","C":2},{"N":"DODGE","C":1213},{"N":"EFFA","C":50},{"N":"EMIS","C":2},{"N":"ENGESA","C":12},{"N":"ENVEMO","C":1},{"N":"FERCAR BUGGY","C":4},{"N":"FERRARI","C":116},{"N":"FIAT","C":49217},{"N":"FIBRAVAN","C":2},{"N":"FORD","C":34977},{"N":"FOTON","C":2},{"N":"FYBER","C":2},{"N":"GEELY","C":20},{"N":"GLASPAC","C":1},{"N":"GMC","C":14},{"N":"GURGEL","C":35},{"N":"HAFEI","C":34},{"N":"HONDA","C":19421},{"N":"HUMMER","C":26},{"N":"HYUNDAI","C":22760},{"N":"INFINITI","C":60},{"N":"INTERNATIONAL","C":1},{"N":"IVECO","C":72},{"N":"JAC","C":602},{"N":"JAGUAR","C":562},{"N":"JEEP","C":6879},{"N":"JINBEI","C":20},{"N":"JPX","C":3},{"N":"KIA","C":6476},{"N":"L AUTOMOBILE","C":1},{"N":"LADA","C":7},{"N":"LAMBORGHINI","C":64},{"N":"LAND ROVER","C":4777},{"N":"LANDWIND","C":1},{"N":"LEXUS","C":177},{"N":"LIFAN","C":620},{"N":"LINCOLN","C":15},{"N":"LOBINI","C":3},{"N":"LOTUS","C":3},{"N":"MAHINDRA","C":8},{"N":"MARCOPOLO","C":14},{"N":"MASERATI","C":74},{"N":"MAZDA","C":22},{"N":"MCLAREN","C":2},{"N":"MENON","C":1},{"N":"MERCEDES-BENZ","C":7751},{"N":"MERCURY","C":5},{"N":"MG","C":19},{"N":"MINI","C":1210},{"N":"MITSUBISHI","C":10466},{"N":"MIURA","C":6},{"N":"MOBBY","C":1},{"N":"MORRIS","C":2},{"N":"MP LAFER","C":21},{"N":"NASH","C":1},{"N":"NISSAN","C":8802},{"N":"OLDSMOBILE","C":2},{"N":"OPEL","C":1},{"N":"PAG","C":1},{"N":"PEUGEOT","C":10472},{"N":"PLYMOUTH","C":2},{"N":"PONTIAC","C":14},{"N":"PORSCHE","C":1282},{"N":"PUMA","C":62},{"N":"RDK","C":2},{"N":"RELY","C":23},{"N":"RENAULT","C":25109},{"N":"ROLLS-ROYCE","C":13},{"N":"SANTA MATILDE","C":5},{"N":"SATURN","C":1},{"N":"SEAT","C":7},{"N":"SELVAGEM","C":1},{"N":"SHELBY","C":6},{"N":"SHINERAY","C":15},{"N":"SHUANGHUAN","C":1},{"N":"SIMCA","C":12},{"N":"SMART","C":200},{"N":"SSANGYONG","C":305},{"N":"SUBARU","C":644},{"N":"SUZUKI","C":1638},{"N":"TAC","C":2},{"N":"TESLA","C":5},{"N":"TOYOTA","C":20028},{"N":"TROLLER","C":464},{"N":"VOLKSWAGEN","C":51967},{"N":"VOLVO","C":2317},{"N":"WAKE","C":3},{"N":"WAY BRASIL","C":2},{"N":"WILLYS","C":99},{"N":"WILLYS OVERLAND","C":8}]}',
			'moto':'{"Principal":[{"N":"BMW","C":1850},{"N":"DAFRA","C":589},{"N":"DUCATI","C":512},{"N":"HARLEY-DAVIDSON","C":2181},{"N":"HONDA","C":11919},{"N":"KASINSKI","C":130},{"N":"KAWASAKI","C":1889},{"N":"SUNDOWN","C":21},{"N":"SUZUKI","C":2557},{"N":"YAMAHA","C":4484}],"Common":[{"N":"AGRALE","C":1},{"N":"AMAZONAS","C":1},{"N":"AME AMAZONAS","C":2},{"N":"APRILIA","C":6},{"N":"BENELLI","C":3},{"N":"BENZHOU","C":9},{"N":"BIMOTA","C":2},{"N":"BMW","C":1850},{"N":"BRANDY","C":2},{"N":"BUELL","C":30},{"N":"BY CRISTO","C":1},{"N":"CAGIVA","C":2},{"N":"CAN-AM","C":32},{"N":"DAFRA","C":589},{"N":"DAYUN","C":2},{"N":"DUCATI","C":512},{"N":"FUN MOTORS","C":5},{"N":"FYM","C":8},{"N":"GARINNI","C":2},{"N":"GAS GAS","C":1},{"N":"GREEN","C":1},{"N":"HAOJUE","C":213},{"N":"HARLEY-DAVIDSON","C":2181},{"N":"HONDA","C":11919},{"N":"HUSQVARNA","C":1},{"N":"INDIAN","C":46},{"N":"IROS","C":2},{"N":"JAWA","C":1},{"N":"KASINSKI","C":130},{"N":"KAWASAKI","C":1889},{"N":"KTM","C":164},{"N":"KYMCO","C":72},{"N":"LAMBRETTA","C":3},{"N":"LINZHI","C":2},{"N":"MALAGUTI","C":2},{"N":"MOTO GUZZI","C":2},{"N":"MOTORINO","C":1},{"N":"MULLER","C":1},{"N":"MV AGUSTA","C":76},{"N":"MVK","C":1},{"N":"MXF","C":1},{"N":"PIAGGIO","C":31},{"N":"POLARIS","C":11},{"N":"REGAL RAPTOR","C":1},{"N":"RIGUETE","C":1},{"N":"ROYAL ENFIELD","C":2},{"N":"SACHS","C":1},{"N":"SANYANG","C":1},{"N":"SHERCO","C":2},{"N":"SHINERAY","C":20},{"N":"SUNDOWN","C":21},{"N":"SUZUKI","C":2557},{"N":"TRAXX","C":18},{"N":"TRICICAR","C":1},{"N":"TRIUMPH","C":767},{"N":"VESPA","C":2},{"N":"WUYANG","C":1},{"N":"YAMAHA","C":4484}]}'
		}
		
		for model_type in brand_list:

			dir_path = os.path.join('scraped-results', model_type)

			if not os.path.exists(dir_path):

				os.makedirs(dir_path)

			brands = json.loads(brand_list[model_type])

			brands = brands['Common'] + brands['Principal']

			brandIdx = 0

			for brand in brands:

				brandIdx += 1

				if brandIdx < 4:

					dir_path = os.path.join('scraped-results', model_type, brand['N'].encode('utf-8').strip())

					if not os.path.exists(dir_path):
						
						os.makedirs(dir_path)

					brand = brand['N'].encode('utf-8').strip().lower().replace(' ', '-')

					### create xlsx file for each brand
					workbook = Workbook()

					main_table_sheet = workbook.active

					main_table_sheet.title = 'Scraped Results'

					for index, header in enumerate(self.main_table_headers):

						main_table_sheet.cell( column = index+1, row = 1, value = header )

					output_file_name = os.path.join(dir_path, 'data.xlsx')

					os.remove(output_file_name) if os.path.exists(output_file_name) else None

					self.workbook_list[brand] = {
						'filename': output_file_name,
						'workbook': workbook,
						'cur_rowIndex': 2
					}
					### create xlsx file for each brand

					search_page_url = ('https://www.webmotors.com.br/%ss/estoque/%s?qt=36' % (model_type, brand))

					yield scrapy.Request(url=search_page_url, callback=self.parse_search_result_page, meta={'page_id': 1, 'search_page_url': search_page_url, 'm_type': model_type, 'dir_path': dir_path, 'brand': brand}, headers=self.headers)

	# parse search result page
	def parse_search_result_page(self, response):

		m_type = response.meta['m_type']

		selector_list = response.xpath('//div[contains(@class, "boxResultado")]/a')

		if len(selector_list) > 0:

			for selector in selector_list:
				try:

					item = {}

					detail_link = selector.xpath('./@href').extract_first()

					item['version_of_year'] = self.validate(selector.xpath('.//span[@class="version"]/text()').extract_first())

					item['financing'] = self.validate(', '.join(selector.xpath('.//div[@class="mrg-left attributes"]/span/text()').extract()))

					item['city'] = self.validate(selector.xpath('.//div[contains(@class, "card-footer")]/span[1]/text()').extract_first())

					item['a_type'] = self.validate(selector.xpath('.//div[contains(@class, "card-footer")]/span[1]/text()').extract_first())

					yield scrapy.Request(url=detail_link, callback=self.parse_detail_page, meta={'item': item, 'm_type': m_type, 'dir_path': response.meta['dir_path'], 'brand': response.meta['brand']}, headers=self.headers)

				except:

					pass


			page_id = response.meta['page_id']
			page_id += 1

			search_page_url = (('%s&p=%s') % (response.meta['search_page_url'], page_id))

			yield scrapy.Request(url=search_page_url, callback=self.parse_search_result_page, meta={'page_id': page_id, 'm_type': m_type, 'search_page_url': response.meta['search_page_url'], 'dir_path': response.meta['dir_path'], 'brand': response.meta['brand']},headers=self.headers)


	# parse detail page
	def parse_detail_page(self, response):

		item = response.meta['item']

		item['make_model'] = self.validate(response.xpath('//span[contains(@class, "makemodel")]/text()').extract_first())

		item['price'] = self.validate(response.xpath('//span[contains(@class, "b__price")]/text()').extract_first())

		item['store_name'] = self.validate(response.xpath('//strong[@class="store-name"]/text()').extract_first())

		item['year'] = self.validate(response.xpath('//div[@class="vehicle-details__main-info"]/table//tr[1]/td[1]/span/text()').extract_first())

		item['mileage'] = self.validate(response.xpath('//div[@class="vehicle-details__main-info"]/table//tr[1]/td[2]/span/text()').extract_first())

		item['exchange'] = self.validate(response.xpath('//div[@class="vehicle-details__main-info"]/table//tr[1]/td[3]/span/text()').extract_first())

		item['fuel'] = self.validate(response.xpath('//div[@class="vehicle-details__main-info"]/table//tr[1]/td[4]/span/text()').extract_first())

		item['color'] = self.validate(response.xpath('//div[@class="vehicle-details__main-info"]/table//tr[2]/td[1]/span/text()').extract_first())

		item['bodywork'] = self.validate(response.xpath('//div[@class="vehicle-details__main-info"]/table//tr[2]/td[2]/span/text()').extract_first())

		item['end_plate'] = self.validate(response.xpath('//div[@class="vehicle-details__main-info"]/table//tr[2]/td[3]/span/text()').extract_first())

		item['items'] = self.validate(', '.join(response.xpath('//div[contains(@class, "size-default")]/div/table//td/text()').extract()))

		item['seller_notes'] = self.validate(response.xpath('//div[contains(@class, "size-default")]//p[@class="info-seller"]/text()').extract_first())

		imgURL_list = response.xpath('//div[contains(@class, "carousel-inner")]//img/@src').extract()

		images = []

		for imgURL in imgURL_list:

			imgURL = imgURL.split('?')[0]

			images.append(self.validate(imgURL))

			# yield scrapy.Request(imgURL, callback=self.image_download, meta={'dir_path': response.meta['dir_path']},headers=self.headers)

		item['images'] = ', '.join(images)

		code = self.validate(response.xpath('//form/input[@name="codigoAnuncio"]/@value').extract_first())

		m_type = response.meta['m_type']

		phone_url = ('https://www.webmotors.com.br/comprar/versomentetelefone?codigoAnuncio=%s&tipoVeiculo=%s' % (code, m_type))

		yield scrapy.Request(url=phone_url, callback=self.parse_phone_number, meta={'item': item, 'dir_path': response.meta['dir_path'], 'brand': response.meta['brand']},headers=self.headers)


	# download images
	def image_download(self, response):

		dir_path = os.path.join(response.meta['dir_path'], 'images')

		if not os.path.exists(dir_path):
			
			os.makedirs(dir_path)
		
		name = dir_path + '/' + response.url.split('/')[-1]

		with open(name, 'wb') as f:

			f.write(response.body)


	# parse phone number
	def parse_phone_number(self, response):
		phone_list = json.loads(response.body_as_unicode().replace("'", '"'))

		phone_numbers = []

		for phone in phone_list:
			phone_numbers.append(self.validate(phone['Telefone']))

		item = response.meta['item']

		item['phone_numbers'] = ', '.join(phone_numbers)

		brand = response.meta['brand']

		main_table_sheet = self.workbook_list[brand]['workbook'].active

		for key, value in item.items():

			main_table_sheet.cell( column = self.main_table_headers.index(key) + 1, row = self.workbook_list[brand]['cur_rowIndex'], value = value)

		self.workbook_list[brand]['cur_rowIndex'] += 1

		self.workbook_list[brand]['workbook'].save( filename = self.workbook_list[brand]['filename'] )

		self.s_count += 1

		print "================================================================"
		print "================================================================"
		print "================================================================"
		print "========================== %s =================" % (self.s_count)
		print "================================================================"
		print item
		print "================================================================"
		print "================================================================"
		print "================================================================"

		# yield item
