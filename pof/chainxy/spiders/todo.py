# from __future__ import unicode_literals
import scrapy
import json
import os
import scrapy
from scrapy.spiders import Spider
from scrapy.http import FormRequest
from scrapy.http import Request
from chainxy.items import ChainItem
from lxml import etree
from lxml import html
import pdb
import time
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook 



class todo(scrapy.Spider):

	name = 'pof'

	domain = 'pof.com'

	history = []

	credentials = {
		'username': 'sm9448',
		'password': 'dhkdwk01'
	}

	xlsx_headers = []

	workbook_list = {}


	def __init__(self):

		with open('require/states.json', 'r') as f:

			self.states = json.load(f)

		with open('require/cities.json', 'r') as f:

			self.cities = json.load(f)

		with open('require/headers.txt', 'r') as f:

			for header in f.readlines():

				self.xlsx_headers.append(self.validate(header))

		None if os.path.exists('scraped-results') else os.makedirs('scraped-results')

		self.driver = webdriver.Chrome("./chromedriver.exe")

	
	def start_requests(self):

		url  = 'https://www.pof.com.br'

		yield scrapy.Request(url=url, callback=self.parse, meta={'state_id': 41}) 


	def parse(self, response):

		### Log In to POF
		self.driver.get('https://www.pof.com.br/inbox.aspx')

		self.driver.find_element_by_id('logincontrol_username').clear()

		self.driver.find_element_by_id('logincontrol_username').send_keys(self.credentials['username'])

		self.driver.find_element_by_id('logincontrol_password').clear()

		self.driver.find_element_by_id('logincontrol_password').send_keys(self.credentials['password'])

		self.driver.find_element_by_id('logincontrol_submitbutton').click()

		self.driver.get('https://www.pof.com.br/pt_basicsearch.aspx')

		for city in self.cities:

			if city['state_id'] not in self.workbook_list:

				workbook = Workbook()

				main_table_sheet = workbook.active

				main_table_sheet.title = 'Scraped Results'

				for index, header in enumerate(self.xlsx_headers):

					main_table_sheet.cell( column = index+1, row = 1, value = header )

				dir_path = os.path.join('scraped-results/profiles')
					
				None if os.path.exists(dir_path) else os.makedirs(dir_path);

				output_file_name = os.path.join(dir_path, 'state_%s.xlsx' % (city['state_id']))

				os.remove(output_file_name) if os.path.exists(output_file_name) else None

				self.workbook_list[city['state_id']] = {
					'filename': output_file_name,
					'workbook': workbook,
					'cur_rowIndex': 2				
				}

			Select(self.driver.find_element_by_id("country")).select_by_visible_text("Brasil")

			self.driver.find_element_by_id("city1").clear()

			self.driver.find_element_by_id("city1").send_keys(city['name'])

			Select(self.driver.find_element_by_id("milesa")).select_by_visible_text("200 miles")

			self.driver.find_element_by_name("cmdSearch").click()

			while True:

				source = self.driver.page_source

				tree = etree.HTML(source)

				for detail_link in tree.xpath('//div[@class="profile"]/a[2]/@href'):

					url = "https://www.pof.com.br/"+detail_link


					if url not in self.history:

						self.history.append(url)

						try:

							yield scrapy.Request(url=url, callback=self.parse_detail_page, meta={'state_id': city['state_id']})

						except:

							pass


				try:

					self.driver.find_element_by_id('basicsearch_nextpage').click()

				except:

					break
	

	def parse_detail_page(self, response):
		profile = {}

		profile['Username'] = self.validate(response.xpath('//span[@id="username"]/text()').extract_first())

		profile['Headline'] = self.validate(response.xpath('//span[@id="headline"]/text()').extract_first())

		for detail in response.xpath('//div[@id="user-details-wrapper"]/div'):

			header = self.validate(' '.join(detail.xpath('./div[1]//text()').extract()))

			data = self.validate(' '.join(detail.xpath('./div[2]//text()').extract()))

			profile[header] = data

		for detail in response.xpath('//div[@class="aligncenter"]/table[@width="85%"]//tr'):

			try:

				header = self.validate(' '.join(detail.xpath('./td[1]//text()').extract()))

				data = self.validate(' '.join(detail.xpath('./td[2]//text()').extract()))

				profile[header] = data

			except:

				print '============ tr tag is empty ==============='

			try:

				header = self.validate(' '.join(detail.xpath('./td[4]//text()').extract()))

				data = self.validate(' '.join(detail.xpath('./td[5]//text()').extract()))

				profile[header] = data

			except:

				print '=========== second language is none ==================='


		profile['Images'] = '\n'.join(response.xpath('//div[@class="image-thumb-wrap"]//img/@src').extract())

		for img_url in response.xpath('//div[@class="image-thumb-wrap"]//img/@src').extract():

			yield scrapy.Request(url=img_url, callback=self.image_download)

		state_id = response.meta['state_id']

		main_table_sheet = self.workbook_list[state_id]['workbook'].active

		for key, value in profile.items():

			if key:

				main_table_sheet.cell( column = self.xlsx_headers.index(key.encode('utf8')) + 1, row = self.workbook_list[state_id]['cur_rowIndex'], value = value)

		self.workbook_list[state_id]['cur_rowIndex'] += 1

		self.workbook_list[state_id]['workbook'].save( filename = self.workbook_list[state_id]['filename'] )

	

	# download images
	def image_download(self, response):

		None if os.path.exists('scraped-results/images/') else os.makedirs('scraped-results/images/')

		name = 'scraped-results/images/' + response.url.split('/')[-1]

		with open(name, 'wb') as f:

			f.write(response.body)

	


	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').strip()

		except:

			pass