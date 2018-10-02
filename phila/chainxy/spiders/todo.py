# from __future__ import unicode_literals
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from lxml import etree

from lxml import html

import pdb

import datetime

import pytz

from openpyxl import Workbook

import random



class phila(scrapy.Spider):

	name = 'phila'

	domain = ''

	history = []

	workbook_list = {}

	headers = [
		'Property Street',
		'Property City',
		'Property State',
		'Property Zipcode',
		'Name of Owner',
		'Owner Street',
		'Owner City',
		'Owner State',
		'Owner Zipcode',
		'2019 Market Value',
		'Sales Price',
		'Sales Date',
		'Description',
		'Condition',
		'Land Area',
		'Improvement Area',
		'Zoning Code'
	]

	street_list = []

	proxy_list = []


	def __init__(self):

		dir_path = 'scraped-result'

		if not os.path.exists(dir_path):

			os.makedirs(dir_path)


		street_list = []

		with open('require/streets.json') as f:

			street_list = json.load(f)

		for street_name in street_list:

			if street_name not in self.street_list:

				self.street_list.append(street_name)


		with open('require/proxy_list.json') as f:

			self.proxy_list = json.load(f)



	def create_file(self, street_name):

		workbook = Workbook()

		main_table_sheet = workbook.active

		main_table_sheet.title = 'Scraped Results'

		for index, header in enumerate(self.headers):

			main_table_sheet.cell( column = index+1, row = 1, value = header )

		output_file_name = 'scraped-result/scraped_data(%s).xlsx' % (street_name)

		os.remove(output_file_name) if os.path.exists(output_file_name) else None

		self.workbook_list[street_name] = {
			'filename': output_file_name,
			'workbook': workbook,
			'cur_rowIndex': 2
		}


	
	def start_requests(self):

		yield scrapy.Request('https://property.phila.gov', callback=self.parse_start)


	def parse_start(self, response):

		for street_name in self.street_list:

			for bn_num in range(1, 100):

				self.create_file(street_name)

				proxy = random.choice(self.proxy_list)

				url  = 'https://api.phila.gov/ais_ps/v1/block/%s %s?include_units=&opa_only=' % (bn_num*100, street_name)

				yield scrapy.Request(url=url, callback=self.parse, meta={'bn_num': bn_num, 'street_name': street_name, 'proxy': 'http://%s' % (proxy)})


	def parse(self, response):

		bn_num = response.meta['bn_num']

		street_name = response.meta['street_name']
			
		try:

			json_data = json.loads(response.body)

			for data in json_data['features']:

				item = {}

				if 'properties' in data:

					propertyInfo = data['properties']

					item['Property Street'] = propertyInfo['opa_address']
					
					item['Property City'] = 'Philadelphia'
					
					item['Property State'] = 'PA'
					
					item['Property Zipcode'] = propertyInfo['zip_code']

					if propertyInfo['zip_4']:

						item['Property Zipcode'] = ('%s - %s') % (item['Property Zipcode'], propertyInfo['zip_4'])

					item['Zoning Code'] = propertyInfo['zoning']

					url = 'https://phl.carto.com/api/v2/sql?q=select+*+from+opa_properties_public+where+++++++++++++++++++parcel_number = \'%s\'' % (propertyInfo['opa_account_num'])

					proxy = random.choice(self.proxy_list)

					yield scrapy.Request(url=url, callback=self.parseOwner, meta={'item': item,'bn_num': bn_num, 'street_name': street_name, 'proxy': 'http://%s' % (proxy)})

		except:

			pass
		


	def parseOwner(self, response):

		try:

			json_data = json.loads(response.body)

			if 'rows' in json_data:

				item = response.meta['item']

				owner = json_data['rows'][0]

				item['Name of Owner'] = owner['owner_1']

				try:

					city_state = owner['mailing_city_state'].split()

					item['Owner Street'] = owner['mailing_street']

					item['Owner State'] = city_state.pop()

					item['Owner City'] = ' '.join(city_state)

					item['Owner Zipcode'] = owner['mailing_zip']

				except:

					item['Owner Street'] = item['Property Street']

					item['Owner State'] = item['Property State']

					item['Owner City'] = item['Property City']

					item['Owner Zipcode'] = item['Property Zipcode']


				item['2019 Market Value'] = owner['market_value']

				item['Sales Price'] = owner['sale_price']

				utc_time = datetime.datetime.strptime(owner['sale_date'], '%Y-%m-%dT%H:%M:%SZ')

				item['Sales Date'] = pytz.utc.localize(utc_time).astimezone(pytz.timezone('America/New_York')).strftime('%Y-%m-%d')

				item['Description'] = owner['building_code_description']

				conditions = {
					'1': 'Other',
					'2': 'New / Rehabbed',
					'3': 'Above Average',
					'4': 'Average',
					'5': 'Below Average',
					'6': 'Vacant',
					'7': 'Sealed / Structurally Compromised'
				}

				item['Condition'] = conditions.get(owner['exterior_condition'], 'Not Applicable')

				item['Land Area'] = round(owner['total_area'])

				item['Improvement Area'] = round(owner['total_livable_area'])


				bn_num = response.meta['bn_num']

				street_name = response.meta['street_name']

				main_table_sheet = self.workbook_list[street_name]['workbook'].active

				for key, value in item.items():

					main_table_sheet.cell( column = self.headers.index(key) + 1, row = self.workbook_list[street_name]['cur_rowIndex'], value = value)

				self.workbook_list[street_name]['cur_rowIndex'] = self.workbook_list[street_name]['cur_rowIndex'] + 1

				self.workbook_list[street_name]['workbook'].save( filename = self.workbook_list[street_name]['filename'] )


				print '============================================================================='
				print '============================================================================='
				print '============================================================================='
				print '============================ %s, %s ================================' % (bn_num, street_name)
				print '============================================================================='
				print '============================================================================='
				print '============================================================================='

		except:

			pass


	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').strip()

		except:

			pass


	def eliminate_space(self, items):

		tmp = []

		for item in items:

			if self.validate(item) != '':

				tmp.append(self.validate(item))

		return tmp