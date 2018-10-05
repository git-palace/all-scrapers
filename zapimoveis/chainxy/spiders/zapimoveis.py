# from __future__ import unicode_literals
# -*- coding: utf-8 -*-
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from lxml import etree

from lxml import html

from openpyxl import Workbook

import pdb



class zapimoveis(scrapy.Spider):

	name = 'zapimoveis'

	domain = 'www.zapimoveis.com.br'

	history = {
		'RENT': [],
		'SALE': []
	}

	headers = {
		'x-requested-with': 'XMLHttpRequest',
		"accept": "*/*",		
		"accept-encoding": "gzip, deflate, br",		
		"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36,",
		"content-type": "application/x-www-form-urlencoded; charset=UTF-8"
	}

	scraped_count = {
		'RENT': 0,
		'SALE': 0
	}

	businessTypeList = ['RENT', 'SALE']

	workbooks= {
		'RENT': [],
		'SALE': []
	}

	sheet_headers = {}

	hashFragments = {}

	fileLimit = 3000


	def __init__(self):	

		headers = [
			'Title',
			'Business Type',
			'Address',
			'Area',
			'Bedrooms',
			'Suites',
			'Parkings',
			'Price',
			'Condo Price',
			'Caracteristicas',
			'Property Type',
			'Phones',
			'Description',
			'Características das Áreas Comuns',
			'VALOR DO M2',
			'ÁREA TOTAL (M2)',
			'Images',
			'Registrant Name'
		]

		for businessType in self.businessTypeList:

			path ='require/params_'+businessType+'.json'

			self.sheet_headers[businessType] = headers

			if businessType == 'RENT':

				self.sheet_headers[businessType] = self.sheet_headers[businessType] + ['VENDA', 'IPTU Price']

			with open(path, 'r') as f:

				self.hashFragments[businessType] = json.load(f)

	


	def start_requests(self):

		url  = 'https://www.zapimoveis.com.br'

		yield scrapy.Request(url=url, callback=self.parse_start) 


	def populateFormData(self, _hashFragment, pageId, businessType):

		hashFragment = {
			"precomaximo": "2147483647",
			"parametrosautosuggest": _hashFragment['parametrosautosuggest'],
			"pagina": pageId,
			"ordem": "Relevancia",
			"paginaOrigem": "ResultadoBusca",
			"semente": "1655658893",
			"formato": "Lista"
		}

		formData = {
			'tipoOferta': '1',
			'paginaAtual': str(pageId),
			'ordenacaoSelecionada': 'Relevancia',
			'pathName': '/%s/imoveis/%s/' % ('venda' if businessType == 'SALE' else 'aluguel', _hashFragment['parametrosautosuggest'][0]['Estado'].lower()),
			'hashFragment': json.dumps(hashFragment),
			'formato': 'Lista'
		}

		return formData


	def parse_start(self, response):

		for businessType in self.businessTypeList:

			for _hashFragment in self.hashFragments[businessType]:

				pageId = 1

				formData = self.populateFormData(_hashFragment, pageId, businessType)

				yield scrapy.FormRequest(url='https://www.zapimoveis.com.br/Busca/RetornarBuscaAssincrona', headers=self.headers, method='POST', formdata=formData, callback=self.parse, meta={'pageId': pageId, '_hashFragment': _hashFragment, 'businessType': businessType})


	def getCaracteristicas(self, _caracteristicas):

		caracteristicas = []

		for caracteristica in _caracteristicas:

			if caracteristica not in caracteristicas:

				caracteristicas.append(caracteristica)

		return caracteristicas


	def parse(self, response):

		_hashFragment = response.meta['_hashFragment']

		pageId = response.meta['pageId'] + 1

		businessType = response.meta['businessType']

		data = json.loads(response.body)

		if 'Resultado' in data and 'Resultado' in data['Resultado'] and len(data['Resultado']['Resultado']):

			for _property in data['Resultado']['Resultado']:

				try:

					if _property['CodImobiliaria'] and _property['CodigoOfertaZAP']:

						propertyID = '%s - %s' % (_property['CodImobiliaria'], _property['CodigoOfertaZAP'])

						if propertyID not in self.history[businessType]:

							self.history[businessType].append(propertyID)

							imgURLs = self.parseImages(_property['Fotos'])

							item = {
								'Business Type': businessType,
								'Address': '%s, %s, %s - %s' % (_property['Endereco'], _property['Bairro'], _property['Cidade'], _property['Estado']),
								'Area': _property['Area'],
								'Bedrooms': _property['QuantidadeQuartos'],
								'Suites': _property['QuantidadeSuites'],
								'Parkings': _property['QuantidadeVagas'],
								'Price': _property['Valor'],
								'Condo Price': _property['PrecoCondominio'],
								'Caracteristicas': ',\n'.join(self.getCaracteristicas(_property['Caracteristicas'])),
								'Property Type': _property['TipoImovel'] if _property['TipoImovel'] else 'APARTAMENTO' + (' COMPRAR' if businessType == 'SALE' else ' ALUGAR'),
								'Phones': '',
								'Description': _property['Observacao'],
								'Características das Áreas Comuns': '',
								'VALOR DO M2': '',
								'ÁREA TOTAL (M2)': '',
								'Images': ',\n'.join(imgURLs),
								'Registrant Name': _property['NomeAnunciante']
							}

							item['Title'] = '%s, %s, %s, %s' % (item['Property Type'], _property['QuantidadeQuartosFormatada'], _property['AreaFormatada'].replace('m<sup>2</sup>', 'M2'), item['Price'])

							if businessType == 'RENT':

								item['IPTU Price'] = _property['ValorIPTU']


							if (self.scraped_count[businessType] >= self.fileLimit):

								self.createExcelFile(businessType)

								self.scraped_count[businessType] = 0

							self.writeExcel(businessType, item)

							# for imgURL in imgURLs:

								# yield scrapy.Request(url=imgURL, callback=self.downloadImage, meta={'businessType': businessType, 'folderIdx': len(self.workbooks[businessType]) + 1} )

				except:

					with open('res.txt', 'wb') as f:

						f.write(response.body)

					pdb.set_trace()

					if _property['CodImobiliaria'] and _property['CodigoOfertaZAP']:

						propertyID = '%s - %s' % (_property['CodImobiliaria'], _property['CodigoOfertaZAP'])

						if propertyID not in self.history[businessType]:

							self.history[businessType].append(propertyID)

							imgURLs = self.parseImages(_property['Fotos'])

							item = {
								'Business Type': businessType,
								'Address': '%s, %s, %s - %s' % (_property['Endereco'], _property['Bairro'], _property['Cidade'], _property['Estado']),
								'Area': _property['Area'],
								'Bedrooms': _property['QuantidadeQuartos'],
								'Suites': _property['QuantidadeSuites'],
								'Parkings': _property['QuantidadeVagas'],
								'Price': _property['Valor'],
								'Condo Price': _property['PrecoCondominio'],
								'Caracteristicas': ',\n'.join(self.getCaracteristicas(_property['Caracteristicas'])),
								'Property Type': _property['TipoImovel'] if _property['TipoImovel'] else 'APARTAMENTO' + (' COMPRAR' if businessType == 'SALE' else ' ALUGAR'),
								'Phones': '',
								'Description': _property['Observacao'],
								'Características das Áreas Comuns': '',
								'VALOR DO M2': '',
								'ÁREA TOTAL (M2)': '',
								'Images': ',\n'.join(imgURLs),
								'Registrant Name': _property['NomeAnunciante']
							}

							item['Title'] = '%s, %s, %s, %s' % (item['Property Type'], _property['QuantidadeQuartosFormatada'], _property['AreaFormatada'].replace('m<sup>2</sup>', 'M2'), item['Price'])

							if businessType == 'RENT':

								item['IPTU Price'] = _property['ValorIPTU']


							if (self.scraped_count[businessType] >= self.fileLimit):

								self.createExcelFile(businessType)

								self.scraped_count[businessType] = 0

							self.writeExcel(businessType, item)


			if 'oldResponse' not in response.meta or response.meta['oldResponse'] != data['Resultado']['Resultado']:

				formData = self.populateFormData(_hashFragment, pageId, businessType)

				yield scrapy.FormRequest(url='https://www.zapimoveis.com.br/Busca/RetornarBuscaAssincrona', headers=self.headers, method='POST', formdata=formData, callback=self.parse, meta={'pageId': pageId, '_hashFragment': _hashFragment, 'businessType': businessType, 'oldResponse': data['Resultado']['Resultado']})

		pass


	def createExcelFile(self, businessType):

		dir_path = 'scraped-results/%s/sheets/%s' % (businessType, len(self.workbooks[businessType]) + 1)

		os.makedirs(dir_path) if not os.path.exists(dir_path) else None


		workbook = Workbook()

		main_table_sheet = workbook.active

		main_table_sheet.title = 'Results'

		for index, header in enumerate(self.sheet_headers[businessType]):

			try:

				main_table_sheet.cell( column = index + 1, row = 1, value = header )

			except:

				pdb.set_trace()

		output_file_name = os.path.join(dir_path, 'results.xlsx')

		os.remove(output_file_name) if os.path.exists(output_file_name) else None

		file = {
			'filename': output_file_name,
			'workbook': workbook,
			'cur_rowIndex': 2
		}

		self.workbooks[businessType].append(file)



	def writeExcel(self, businessType, item):

		self.createExcelFile(businessType) if not len(self.workbooks[businessType]) else None

		file = self.workbooks[businessType][-1]

		main_table_sheet = file['workbook'].active

		write_flag = True

		for key, value in item.items():

			try:

				main_table_sheet.cell( column = self.sheet_headers[businessType].index(key) + 1, row = file['cur_rowIndex'], value = value)

			except:

				write_flag = False

		if write_flag:

			self.workbooks[businessType][-1]['cur_rowIndex'] += 1

			file['workbook'].save( filename = file['filename'] )

			print '========================================================='
			print '======================= Scraped Count (%s) = %s =====================' % (businessType, (len(self.workbooks[businessType]) - 1) * self.fileLimit + self.scraped_count[businessType])
			print '========================================================='


		self.scraped_count[businessType] += 1


	def parseImages(self, photos):

		imgURLs = []

		for photo in photos:

			imgURL = photo['UrlImagemTamanhoG'].replace('https://imagens.zapcorp.com.br', 'https://zapangelo.zapcorp.com.br/fit-in/620x372/smart/filters:quality(90):format(webp)')

			imgURL = imgURL.replace('_g.', '_raw.')

			imgURLs.append(imgURL)

		return imgURLs



	def downloadImage(self, response):

		dir_path = 'scraped-results/%s/images/%s' % (response.meta['businessType'], response.meta['folderIdx'])

		os.makedirs(dir_path) if not os.path.exists(dir_path) else None
		
		name = dir_path + '/' + response.url.split('/')[-1]

		with open(name, 'wb') as f:

			f.write(response.body)