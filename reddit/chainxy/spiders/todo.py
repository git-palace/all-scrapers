# from __future__ import unicode_literals
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from lxml import etree

from lxml import html

from datetime import datetime

import pdb

from openpyxl import Workbook


class todo(scrapy.Spider):

	name = 'reddit'

	domain = ''

	history = []

	posts = []

	headers = [
		'Username',
		'Created at',
		'Comment Content',
		'Comment Link',
		'Post Title',
		'Post Link'
	]

	cur_index = 2


	def __init__(self):

		self.workbook = Workbook()

		main_table_sheet = self.workbook.active

		main_table_sheet.title = 'Reddit Comments'

		for index, header in enumerate(self.headers):

			main_table_sheet.cell( column = index+1, row = 1, value = header )

		os.remove('results.xlsx') if os.path.exists('results.xlsx') else None

		self.workbook.save('results.xlsx')

	
	def start_requests(self):

		yield scrapy.Request(url='https://www.reddit.com/r/aws.json', callback=self.parse_posts)



	def parse_posts(self, response):

		result = json.loads(response.body)

		go_next_page = True

		try:
			
			for post in result['data']['children']:

				created_at = datetime.utcfromtimestamp(post['data']['created_utc'])

				if (datetime.utcnow() - created_at).days > 90:

					go_next_page = False

					break

				item = {}

				item['Post Title'] = post['data']['title']
				item['Post Link'] = 'https://www.reddit.com' + post['data']['permalink']

				yield scrapy.Request(url='https://www.reddit.com/r/aws/comments/%s.json' % post['data']['id'], callback=self.parse_comments, meta={'item': item})

		except:

			pdb.set_trace()
		
		if go_next_page and ('after' in result['data']) and result['data']['after']:

			posts_url  = 'https://www.reddit.com/r/aws.json?after=' + result['data']['after']
			
			yield scrapy.Request(url=posts_url, callback=self.parse_posts)



	def parse_comments(self, response):

		comments = json.loads(response.body)

		try:

			for comment in comments[1:]:

				self.check_replies(comment, response.meta['item'])
			
				self.workbook.save('results.xlsx')

		except:

			pdb.set_trace()
	


	def check_replies(self, comment, o_item):

		try:

			for c_comment in comment['data']['children']:

				item = o_item

				item['Username'] = c_comment['data']['author']

				item['Created at'] = datetime.utcfromtimestamp(c_comment['data']['created_utc']).strftime('%Y-%m-%d %H:%M:%S')

				item['Comment Content'] = c_comment['data']['body']

				item['Comment Link'] = 'https://www.reddit.com' + c_comment['data']['permalink']

				main_table_sheet = self.workbook.active

				for key, value in item.items():

					main_table_sheet.cell( column = self.headers.index(key) + 1, row = self.cur_index, value = value)
				
				print '=================================================='
				print '=================== %s ==========================' % (self.cur_index - 1)
				print '=================================================='
				self.cur_index = self.cur_index + 1

				if c_comment['data']['replies']:

					self.check_replies(c_comment['data']['replies'], o_item)
				
		except:

			pass